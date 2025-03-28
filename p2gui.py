import tkinter as tk
from tkinter import filedialog, messagebox
from docx import Document
from openpyxl import load_workbook
import xlsxwriter

def find_topword(tokens, freq_table):
    """
    Given a list of tokens and a frequency table, return the token
    with the highest frequency. If none are in the table, return None.
    """
    topword = None
    top_freq = 0
    for token in tokens:
        clean_token = token.strip()
        if clean_token in freq_table and freq_table[clean_token] > top_freq:
            topword = clean_token
            top_freq = freq_table[clean_token]
    return topword

def process_files(input_file, freq_file):
    """
    Contains the logic previously in your `main()` function:
      - Reads the .docx file
      - Reads the frequency table from Excel
      - Writes out referenceTable.xlsx
    """
    # Read DOCX file
    with open(input_file, "rb") as docx_file:
        doc = Document(docx_file)
        unitlist = [paragraph.text for paragraph in doc.paragraphs]

    wb = load_workbook(freq_file, read_only=True, data_only=True)
    ws = wb.active
    freq_table = {}
    for row in ws.iter_rows(values_only=True):
        # Make sure the row has at least two columns with data
        if row[0] is not None and row[1] is not None:
            freq_table[row[0]] = row[1]


    # Create output workbook and worksheet
    workbook = xlsxwriter.Workbook("referenceTable.xlsx")
    worksheet = workbook.add_worksheet()
    underline = workbook.add_format({"underline": True})

    row_index = 0
    for line in unitlist:
        tokens = line.split(",")
        # Find most frequent token
        topword = find_topword(tokens, freq_table)

        # If no token is found in freq_table, just write them all sorted (no underlines)
        if not topword:
            sorted_line = ",".join(sorted(tokens, key=str.upper))
            worksheet.write(row_index, 0, sorted_line)
            row_index += 1
            continue

        # If there is only one token in the line, underline it and write twice
        if len(tokens) == 1:
            worksheet.write(row_index, 0, topword, underline)
            worksheet.write(row_index, 1, topword)
            row_index += 1
            continue

        # Otherwise, remove the topword, sort the rest, and write
        clean_tokens = [t.strip() for t in tokens]
        clean_tokens.remove(topword)
        clean_tokens.sort(key=str.upper)

        # Write topword underlined followed by the rest
        rest_string = ",".join(clean_tokens)
        worksheet.write_rich_string(row_index, 0, underline, topword, ",", rest_string)

        # Write topword again (unformatted) in the second column
        worksheet.write(row_index, 1, topword)
        row_index += 1

    workbook.close()

def select_docx():
    """
    Opens a file dialog to select the input .docx file.
    """
    file_path = filedialog.askopenfilename(
        title="Select DOCX File",
        filetypes=[("DOCX Files", "*.docx"), ("All Files", "*.*")]
    )
    if file_path:
        docx_entry.delete(0, tk.END)
        docx_entry.insert(0, file_path)

def select_excel():
    """
    Opens a file dialog to select the Excel file containing the frequency table.
    """
    file_path = filedialog.askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx *.xls"), ("All Files", "*.*")]
    )
    if file_path:
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, file_path)

def run_processing():
    """
    Validates inputs and calls the process_files function.
    """
    docx_file = docx_entry.get().strip()
    excel_file = excel_entry.get().strip()

    if not docx_file or not excel_file:
        messagebox.showerror("Error", "Please select both a DOCX file and an Excel file.")
        return

    try:
        process_files(docx_file, excel_file)
        messagebox.showinfo(
            "Completed",
            "referenceTable.xlsx has been created successfully!"
        )
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred:\n{e}")

# -------------------------
# Tkinter GUI Setup
# -------------------------
root = tk.Tk()
root.title("DOCX Processing to referenceTable")

# DOCX file
docx_frame = tk.Frame(root)
docx_frame.pack(padx=10, pady=5, fill="x")

docx_label = tk.Label(docx_frame, text="Input DOCX:")
docx_label.pack(side="left")

docx_entry = tk.Entry(docx_frame, width=50)
docx_entry.pack(side="left", padx=5)

docx_button = tk.Button(docx_frame, text="Browse", command=select_docx)
docx_button.pack(side="left")

# Excel file
excel_frame = tk.Frame(root)
excel_frame.pack(padx=10, pady=5, fill="x")

excel_label = tk.Label(excel_frame, text="Excel File:")
excel_label.pack(side="left")

excel_entry = tk.Entry(excel_frame, width=50)
excel_entry.pack(side="left", padx=5)

excel_button = tk.Button(excel_frame, text="Browse", command=select_excel)
excel_button.pack(side="left")

# Run button
run_button = tk.Button(root, text="Process", command=run_processing)
run_button.pack(padx=10, pady=15)

root.mainloop()
